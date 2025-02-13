import os
from datetime import datetime, timedelta
from openpyxl import load_workbook

def write_emails_to_file(emails):
    with open('expiring_emails.txt', 'w') as file:
        for email in emails:
            file.write(email + '\n')    #This format is convinient for gmail accounts, not for Outlook

def main():
    expiring_emails = []

    # Read the Excel
    workbook = load_workbook('user_info.xlsx')
    sheet = workbook.active
    
    # Check the content of the Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, email, expiration_date = row

        reminder_date = expiration_date - timedelta(days=2)
        #print(reminder_date)

        current_date = datetime.now().date()
        #print(current_date)

        # Final Check and message
        if current_date > reminder_date.date() or current_date == reminder_date.date():
            expiring_emails.append(email)
            #print(f"Reminder: {name} subscription will expire on {expiration_date.strftime('%Y-%m-%d')}. Send a message notification in the following email address: {email}")
            #print("")

    write_emails_to_file(expiring_emails)
    print("Emails of expiring users have been written to 'expiring_emails.txt'.")            

if __name__ == "__main__":
    main()
